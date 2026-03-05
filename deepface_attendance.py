import cv2
import os
from deepface import DeepFace
import openpyxl
import time

# Initialize Excel for attendance tracking
def initialize_excel(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Date", "Time", "Status"])
        workbook.save(filename)

def save_to_excel(filename, row):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Attendance"]
    sheet.append(row)
    workbook.save(filename)

# Initialize attendance Excel file
attendance_filename = "attendance.xlsx"
initialize_excel(attendance_filename)

# Path to store the images for reference
reference_images_path = r"C:\D\KHOALUAN\face-recognition-attendance-system\Images"
# Ensure that the reference images folder exists
if not os.path.exists(reference_images_path):
    os.makedirs(reference_images_path)
    print(f"Created Images folder at: {reference_images_path}")
    print("Please add reference face images to the Images folder and restart.")
    exit()

# Load reference images and names from the specified folder
reference_images = {}
for file in os.listdir(reference_images_path):
    if file.endswith(".jpg") or file.endswith(".png"):
        name = os.path.splitext(file)[0].replace("_", " ")
        reference_images[name] = os.path.join(reference_images_path, file)

# Track attendance dictionary
attendance_dict = {}

# Start video capture
cap = cv2.VideoCapture(0)
print("Press 'q' to exit.")

while True:
    ret, frame = cap.read()
    if not ret:
        print("Error: Could not access the camera.")
        break

    # Resize frame for faster processing
    resized_frame = cv2.resize(frame, (640, 480))

    try:
        # Perform face recognition using DeepFace
        results = DeepFace.find(
            img_path=resized_frame,
            db_path=reference_images_path,
            enforce_detection=False,
            detector_backend="opencv"  # Options: opencv, mtcnn, dlib, retinaface, etc.
        )

        if len(results) > 0:
            # Get the first match
            match = results[0].iloc[0]
            name = os.path.basename(match['identity']).split('.')[0].replace("_", " ")

            # Record attendance if not already marked
            if name not in attendance_dict:
                attendance_dict[name] = True
                timestamp = time.strftime("%Y-%m-%d %H:%M:%S").split()
                save_to_excel(attendance_filename, [name, timestamp[0], timestamp[1], "Present"])
                print(f"Marked attendance for: {name}")

            # Display name on the frame
            cv2.putText(frame, name, (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

    except Exception as e:
        print(f"Error: {e}")
        cv2.putText(frame, "Unknown", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

    # Show the video frame
    cv2.imshow("DeepFace Attendance System", frame)

    # Exit on pressing 'q'
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
